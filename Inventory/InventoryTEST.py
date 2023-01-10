import pandas as pd
import openpyxl
from openpyxl import load_workbook

pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

# Select excel workbook
wb = load_workbook(filename='Inventory.xlsx')
sheet = wb.active
df = pd.read_excel('Inventory.xlsx')
df = pd.DataFrame(df, columns = ["Name","Location","Part Number","Quantity","Decryption","Note"])
# View Dataframe
def refreshDataFrame():
    global df
    global sheet
    global wb
    df = pd.read_excel('Inventory.xlsx')
    df = pd.DataFrame(df, columns=["Name", "Location", "Part Number", "Quantity", "Decryption", "Note"])
    print(df)

# sheet['A1'] = 1
# sheet.cell(row=2, column=2).value = 2

def addProduct():
    global df
    global sheet
    global wb
    rows = df.shape[0]
    while True:
        values = []
        user = (input("Please enter the Name: "))
        sheet.cell(row=rows + 2, column=1).value = user
        values.append(user)
        user = (input("Please enter the Location: "))
        sheet.cell(row=rows + 2, column=2).value = user
        values.append(user)
        user = int(input("Please enter the Part Number: "))
        sheet.cell(row=rows + 2, column=3).value = user
        values.append(user)
        user = int(input("Please enter the Quantity: "))
        sheet.cell(row=rows + 2, column=4).value = user
        values.append(user)
        user = int(input("Please enter the Decryption: "))
        sheet.cell(row=rows + 2, column=5).value = user
        values.append(user)
        user = (input("Please enter the Note: "))
        sheet.cell(row=rows + 2, column=7).value = user
        values.append(user)
        cols = ["Name","Location","Part Number","Quantity","Decryption","Note"]
        for i in range(1):
            print('{:<25} {:<25} {:<25} {:<25} {:<25} {:<25}'.format(*cols))
        for i in range(1):
            print('{:<25} {:<25} {:<25} {:<25} {:<25} {:<25}'.format(*values))
        verify = input("Is this information correct? (Y/N)")
        if verify.capitalize() == "Y":
            wb.save('Inventory.xlsx')
            print()
            break
        else:
            continue

def deleteProduct(partnum):
    global df
    global sheet
    global wb
    value = findPartNumber(partnum)
    sheet.delete_rows(value[0] + 2)
    wb.save('Inventory.xlsx')

def changeName(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = (input("Please enter the NEW name you would like to change it to?"))
        sheet.cell(row=value[0] + 2, column=1).value = user
        wb.save('Inventory.xlsx')
        print()
        break

def changeLocation(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = (input("Please enter the NEW location you would like to change it to?"))
        sheet.cell(row=value[0] + 2, column=2).value = user
        wb.save('Inventory.xlsx')
        print()
        break

def findPartNumber(num):
    global df
    global sheet
    global wb
    row = df[df['Part Number'] == num].index
    # List of rows with that same part number
    value = row.tolist()
    # print(row.tolist())
    # print(df.iloc[row])
    if not value:
        print("There is no such product. Please try again.")
        return False
    return value

def changePartNumber(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = int(input("Please enter the NEW part number you would like to change it to?"))
        sheet.cell(row=value[0]+2, column=3).value = user
        wb.save('Inventory.xlsx')
        print()
        break

def changeQuantity(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = int(input("Please enter the NEW quantity you would like to change it to?"))
        sheet.cell(row=value[0] + 2, column=4).value = user
        wb.save('Inventory.xlsx')
        print()
        break

def changeDecryption(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = int(input("Please enter the NEW decryption you would like to change it to?"))
        sheet.cell(row=value[0] + 2, column=5).value = user
        wb.save('Inventory.xlsx')
        print()
        break

def changeNote(partnum):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        # print(value)
        user = (input("Please enter the NEW note you would like to change it to?"))
        sheet.cell(row=value[0] + 2, column=7).value = user
        wb.save('Inventory.xlsx')
        print()
        break

wb.save('Inventory.xlsx')