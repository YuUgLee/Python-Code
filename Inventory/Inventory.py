import pandas as pd
import openpyxl
from openpyxl import load_workbook
# Select excel workbook
wb = load_workbook(filename='Inventory.xlsx')
sheet = wb.active
df = pd.read_excel('Inventory.xlsx')
df = pd.DataFrame(df, columns = ["Name","Location","Part Number","Quantity","Decryption","Note"])
# View Dataframe
def viewDataFrame():
    df = pd.read_excel('Inventory.xlsx')
    df = pd.DataFrame(df, columns=["Name", "Location", "Part Number", "Quantity", "Decryption", "Note"])
    print(df)

# sheet['A1'] = 1
# sheet.cell(row=2, column=2).value = 2



def changeName(row,name):
    sheet.cell(row=row+2, column=1).value = name
    wb.save('Inventory.xlsx')
    print()

def changeLocation(row,name):
    sheet.cell(row=row+2, column=2).value = name
    wb.save('Inventory.xlsx')
    print()

def findPartNumber(num):
    row = df[df['Part Number'] == num].index
    # List of rows with that same part number
    print(row.tolist())
    print(df.iloc[row])
    print()

def changePartNumber(row,name):
    sheet.cell(row=row+2, column=3).value = name
    wb.save('Inventory.xlsx')
    print()

def changeQuantity(row,name):
    sheet.cell(row=row+2, column=4).value = name
    wb.save('Inventory.xlsx')
    print()

def changeDecryption(row,name):
    sheet.cell(row=row+2, column=5).value = name
    wb.save('Inventory.xlsx')
    print()

def changeNote(row,name):
    sheet.cell(row=row+2, column=7).value = name
    wb.save('Inventory.xlsx')
    print()

wb.save('Inventory.xlsx')