import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

cwd = os.getcwd()

#Dataframe settings
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

# Initialize Dataframe
df = pd.read_excel(r'%s\Inventory Files\Inventory.xlsx' %cwd)
df = pd.DataFrame(df, columns = ["Name","Location","Part Number","Decryption","Serial Number","Quantity","Note"])
df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)
df['Part Number'] = df['Part Number'].astype(str)
df['Serial Number'] = df['Serial Number'].astype(str)

# Select excel workbook
wb = load_workbook(filename=r'%s\Inventory Files\Inventory.xlsx' %cwd)
sheet = wb.active
wb.active = 1
# View Dataframe
def refreshDataFrame():
    global df
    global sheet
    global wb
    df = pd.read_excel(r'%s\Inventory Files\Inventory.xlsx' %cwd)
    df = pd.DataFrame(df, columns= ["Name","Location","Part Number","Decryption","Serial Number","Quantity","Note"])

def addProduct(name,location,partnum,qty,decrypt,note,serial,label7):
    global df
    global sheet
    global wb
    rows = df.shape[0]
    sheet.cell(row=rows + 2, column=1).value = name
    sheet.cell(row=rows + 2, column=2).value = location
    sheet.cell(row=rows + 2, column=3).value = partnum
    sheet.cell(row=rows + 2, column=4).value = qty
    sheet.cell(row=rows + 2, column=5).value = decrypt
    sheet.cell(row=rows + 2, column=6).value = serial
    sheet.cell(row=rows + 2, column=8).value = note
    value = findPartNumber(partnum)
    if len(df[df['Part Number'] == partnum].index.tolist()) > 0:
        label7.configure(text="Unable to add product!\nThis Part Number already exists")
        return False
    if len(df[df['Serial Number'] == serial].index.tolist()) > 0:
        label7.configure(text="Unable to add product!\nThis Serial Number already exists")
        return False
    if name == "":
        label7.configure(text="Unable to add product!\nPlease enter a name")
        return False
    else:
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        label7.configure(text=name+" was added to your Inventory!")
        refreshDataFrame()
        df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index = False)

def deleteProduct(partnum, new, label2):
    global df
    global sheet
    global wb
    if str(partnum) == "":
        return
    elif findPartNumber(str(partnum)) == False:
        label2.configure(text = "Part number could not be found. Please try again.")
    else:
        value = findPartNumber(str(partnum))
        sheet.delete_rows(value[0] + 2)
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        label2.configure(text = str(partnum).upper() + " has been deleted!")
        refreshDataFrame()
        df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeName(partnum, user):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=1).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeLocation(partnum, user):
    global df
    global sheet
    global wb
    if user =="nan":
        user=""
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=2).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def findPartNumber(num):
    global df
    global sheet
    global wb

    row = df[df['Part Number'].astype(str) == num.upper()].index
    # List of rows with that same part number
    value = row.tolist()
    # print(row.tolist())
    # print(df.iloc[row])
    if not value:
        # print("There is no such product. Please try again.")
        return False
    return value

def findPartEdit(num):
    global df
    global sheet
    global wb
    row = df[df['Part Number'] == num].index
    # List of rows with that same part number
    value = row.tolist()
    # print(row.tolist())
    # print(df.iloc[row])
    if not value:
        # print("There is no such product. Please try again.")
        return False
    return value

def changePartNumber(partnum, user):
    global df
    global sheet
    global wb
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0]+2, column=3).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeQuantity(partnum, user):
    global df
    global sheet
    global wb

    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=4).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeDecryption(partnum,user):
    global df
    global sheet
    global wb
    if user =="nan":
        user=""
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=5).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeSerial(partnum,user):
    global df
    global sheet
    global wb
    if user =="nan":
        user=""
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=6).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)

def changeNote(partnum,user):
    global df
    global sheet
    global wb
    if user =="nan":
        user=""
    while True:
        value = findPartNumber(partnum)
        sheet.cell(row=value[0] + 2, column=8).value = user
        wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
        break
    refreshDataFrame()
    df.to_excel(r"%s\Inventory Files\Backup\Inventory.xlsx" % (cwd), index=False)
wb.save(r'%s\Inventory Files\Inventory.xlsx' %cwd)
